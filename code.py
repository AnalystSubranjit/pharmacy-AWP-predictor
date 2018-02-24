import sys
import matplotlib.pyplot as plt
import numpy as np
from sklearn import datasets, linear_model
from sklearn.metrics import mean_squared_error, r2_score
import random
from openpyxl import Workbook, load_workbook

def outputToExcel(fileName, list1, list2):
    book = Workbook()
    ws = book.active
    for i in range(1,len(list1)+1):
        ws['A' + str(i)] = list1[i][0]
        ws['B' + str(i)] = list2[i][0]
        if i % 10000 == 0:
            print(str(i*2) + "%")
    book.save(fileName)

def parse(string):
    newString = ""
    ignore = set('"$()%, ')
    for i in string:
        if i not in ignore:
            newString += i
    return newString.split("\t")

if __name__ == "__main__":
    # Read file
    fileName = "pharmacydata.txt"
    fileReader = open(fileName,"r")
    lines = fileReader.readlines()

    totalSales = []
    acq = []
    awp = []
    randomizedLines = lines[1:]
    random.shuffle(randomizedLines)
    
    # Fetch data that fall within certain criteria
    for line in randomizedLines:
        d,ndc,desc,qty,p,s,c,sales,acqCosts,profit,margin,dawcode,manu,acqUnit,price= parse(line)
        sales = float(sales.strip())
        acqCosts = float(acqCosts.strip())
        margin = float(margin.strip())
        price = float(price.strip())
        if 400 > margin > 0 and price < 5000:
            totalSales.append([sales])
            acq.append([acqCosts])
            awp.append([price])

    # Split data into training data and testing data
    x = acq
    y = awp
    trainingSize = len(lines) // 5
    xTrain = x[:trainingSize]
    xTest = x[trainingSize:]
    yTrain = y[:trainingSize]
    yTest = y[trainingSize:]
    
    # Create linear regression model
    model = linear_model.LinearRegression()

    # Train model
    model.fit(xTrain, yTrain)
    model.score(xTrain, yTrain)
    
    # Make predictions using testing set
    predictions = model.predict(xTest)
    
    print("Coefficients: ", model.coef_)
    print("Intercept: ", model.intercept_)
    print("Regression Equation: y = %.4fx+%.4f" %(model.coef_, model.intercept_))
    print("Mean squared error: %.2f" % mean_squared_error(yTest, predictions))
    print('Variance score: %.2f' % r2_score(yTest, predictions))
    
    '''
    # Plot outputs
    plt.scatter(xTest, yTest, color='black')
    plt.plot(xTest, predictions, color='blue', linewidth=3)
    plt.xticks(())
    plt.yticks(())
    plt.show()
    '''
    
    while True:
        inputVal = raw_input("Enter acquisition costs to determine AWP or press Enter to exit:")
        if len(inputVal) == 0:
            break
        try:
            calculatedAWP = model.coef_ * float(inputVal) + model.intercept_
            print("Predicted AWP is : $%.2f" %(calculatedAWP))
        except Exception:
            print("Error:", Exception)
        
    # outputToExcel(newFile, acq, awp)
    # newFile = "results.csv"
    # fileWriter = open(newFile,"w")
    # for i in range(len(x)):
    #     toWrite = str(x[i][0]) + "," + str(y[i][0]) + "\n"
    #     fileWriter.write(toWrite)